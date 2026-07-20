# -*- coding: utf-8 -*-
"""Tạo file Excel biểu mẫu thu thập góp ý cải tiến ERP cho từng phòng ban."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/vunam/Downloads/koola/ERP/ERP_V6/Gop_y_cai_tien_ERP.xlsx"

# ---- Style palette ----
NAVY = "1F2937"
BLUE = "2563EB"
LIGHT = "EEF2FF"
GREY = "F3F4F6"
WHITE = "FFFFFF"
BORDER_CLR = "D1D5DB"

thin = Side(style="thin", color=BORDER_CLR)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

title_font = Font(name="Calibri", size=15, bold=True, color=WHITE)
sub_font = Font(name="Calibri", size=10, italic=True, color="E5E7EB")
head_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
cell_font = Font(name="Calibri", size=10, color="111827")
example_font = Font(name="Calibri", size=10, italic=True, color="6B7280")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Cột: (tiêu đề, độ rộng)
COLUMNS = [
    ("STT", 6),
    ("Chức năng / màn hình liên quan", 26),
    ("Vấn đề / hạn chế gặp phải", 34),
    ("Đề xuất cải tiến", 36),
    ("Lợi ích mong đợi", 28),
    ("Mức ưu tiên", 13),
    ("Độ khó (ước tính)", 14),
    ("Người đề xuất", 16),
    ("Ngày đề xuất", 13),
    ("Ghi chú", 22),
]

PRIORITY = '"Cao,Trung bình,Thấp"'
DIFFICULTY = '"Dễ,Trung bình,Khó,Chưa rõ"'

# Bộ cột riêng cho sheet chức năng chung: có thêm cột "Người yêu cầu".
# Giữ Mức ưu tiên ở cột 6 và Độ khó ở cột 7 để dropdown hoạt động đúng.
COLUMNS_COMMON = [
    ("STT", 6),
    ("Chức năng chung / màn hình", 26),
    ("Vấn đề / hạn chế gặp phải", 34),
    ("Đề xuất cải tiến", 36),
    ("Lợi ích mong đợi", 26),
    ("Mức ưu tiên", 13),
    ("Độ khó (ước tính)", 14),
    ("Người yêu cầu", 16),
    ("Bộ phận", 16),
    ("Ngày đề xuất", 13),
    ("Ghi chú", 22),
]

# Mỗi phòng ban: tên sheet, mô tả module, danh sách ví dụ mẫu (điền sẵn để gợi ý)
# Ví dụ được viết dựa trên các màn hình/chức năng thực tế của hệ thống.
DEPARTMENTS = [
    {
        "sheet": "Chức năng chung",
        "desc": "Tính năng dùng chung toàn hệ thống: đăng nhập, thông báo, tìm kiếm, trợ lý AI, giao diện/tablet, xuất file, lịch sử thao tác",
        "columns": COLUMNS_COMMON,
        "examples": [
            ["Thông báo (Chuông thông báo)",
             "Nhận quá nhiều thông báo, khó lọc loại cần quan tâm",
             "Cho phép cài đặt loại thông báo muốn nhận + lọc theo nhóm",
             "Không bỏ sót việc quan trọng", "Cao", "Trung bình", "", ""],
            ["Đăng nhập / Đổi mật khẩu",
             "Quên mật khẩu phải nhờ admin reset, mất thời gian",
             "Thêm tự đặt lại mật khẩu qua email/OTP",
             "Chủ động, giảm việc cho admin", "Trung bình", "Trung bình", "", ""],
            ["Tìm kiếm chung",
             "Mỗi màn hình tìm kiếm một kiểu, chưa có tìm nhanh toàn hệ thống",
             "Thêm ô tìm kiếm chung (nhân viên, đơn hàng, phiếu...) ở thanh trên",
             "Tìm nhanh, đỡ chuyển qua lại", "Trung bình", "Khó", "", ""],
            ["Giao diện trên tablet",
             "Một số bảng khó thao tác trên màn hình tablet ngoài xưởng",
             "Chuẩn hóa giao diện cảm ứng: nút to, cột gọn cho tablet",
             "Nhập liệu tại chỗ thuận tiện", "Cao", "Trung bình", "", ""],
            ["Xuất file (Excel/PDF)",
             "Nhiều bảng chưa có nút xuất Excel để báo cáo",
             "Bổ sung nút xuất Excel/PDF thống nhất cho các bảng danh sách",
             "Làm báo cáo nhanh hơn", "Trung bình", "Dễ", "", ""],
            ["Trợ lý AI (Chat)",
             "Chưa rõ trợ lý AI hỗ trợ được những gì",
             "Thêm gợi ý câu hỏi mẫu + hướng dẫn ngắn khi mở trợ lý",
             "Tận dụng trợ lý tốt hơn", "Thấp", "Dễ", "", ""],
        ],
    },
    {
        "sheet": "Chung - Nhân sự",
        "desc": "Nhân sự, chấm công, bảng lương, đánh giá nhân viên, tài sản, nghỉ phép, tăng ca",
        "examples": [
            ["Bảng chấm công tháng (Chấm công)",
             "Lưới chấm công tháng tải chậm khi phòng đông người, khó thao tác trên tablet",
             "Tối ưu tải theo trang/nhóm, cho phép lọc theo tổ/ca trước khi hiển thị",
             "Chấm công nhanh, giảm thời gian chờ", "Cao", "Trung bình"],
            ["Bảng lương (Lương)",
             "Chưa xuất được phiếu lương từng nhân viên gửi qua email/thông báo",
             "Thêm nút xuất phiếu lương PDF + gửi thông báo cho nhân viên",
             "Nhân viên tự xem lương, giảm hỏi đáp thủ công", "Trung bình", "Trung bình"],
            ["Nghỉ phép (Đơn nghỉ phép)",
             "Không thấy số ngày phép còn lại khi tạo đơn",
             "Hiển thị số phép năm còn lại ngay trên form tạo đơn",
             "Tránh nộp đơn vượt phép, giảm từ chối", "Trung bình", "Dễ"],
        ],
    },
    {
        "sheet": "Tổng hợp",
        "desc": "Phòng giá thành (báo giá, tính giá), Phòng chăm sóc đối tác",
        "examples": [
            ["Tính giá thành (Báo giá)",
             "Nhập lại thông số lặp lại cho các báo giá tương tự tốn thời gian",
             "Cho phép sao chép báo giá cũ làm mẫu",
             "Tạo báo giá nhanh hơn", "Trung bình", "Dễ"],
            ["Quản lý báo giá",
             "Khó theo dõi trạng thái báo giá đã gửi/chờ duyệt",
             "Thêm bộ lọc + màu trạng thái trong danh sách báo giá",
             "Nắm nhanh tình trạng, không bỏ sót", "Trung bình", "Dễ"],
        ],
    },
    {
        "sheet": "Chất lượng",
        "desc": "CL nhân sự (đánh giá năng lực), CL quy trình (kiểm tra nội bộ, quy trình)",
        "examples": [
            ["Kiểm tra nội bộ",
             "Ghi nhận lỗi phải nhập tay nhiều, chưa có mẫu lỗi thường gặp",
             "Thêm danh mục lỗi mẫu để chọn nhanh khi ghi nhận",
             "Ghi nhận nhanh, thống kê lỗi chuẩn hơn", "Cao", "Trung bình"],
            ["Đánh giá chất lượng nhân sự",
             "Khó so sánh kết quả đánh giá giữa các kỳ",
             "Thêm biểu đồ xu hướng điểm theo kỳ cho từng nhân viên",
             "Theo dõi tiến bộ rõ ràng", "Trung bình", "Trung bình"],
        ],
    },
    {
        "sheet": "Kinh doanh",
        "desc": "KD Quốc tế, KD Nội địa (khách hàng, đơn hàng, công nợ, phản hồi KH)",
        "examples": [
            ["Quản lý đơn hàng",
             "Chưa cảnh báo đơn sắp đến hạn giao",
             "Thêm nhắc nhở/thông báo đơn hàng gần đến hạn giao",
             "Giao hàng đúng hạn, tăng hài lòng KH", "Cao", "Trung bình"],
            ["Công nợ",
             "Khó lọc nhanh các khoản nợ quá hạn theo khách hàng",
             "Thêm bộ lọc quá hạn + tổng hợp nợ theo khách hàng",
             "Thu hồi công nợ hiệu quả hơn", "Cao", "Trung bình"],
            ["Phản hồi khách hàng",
             "Không rõ phản hồi nào đã xử lý/chưa xử lý",
             "Thêm trạng thái xử lý + người phụ trách cho mỗi phản hồi",
             "Không bỏ sót phản hồi KH", "Trung bình", "Dễ"],
        ],
    },
    {
        "sheet": "Kế toán",
        "desc": "KT Hành chính, KT Thuế (hóa đơn, báo cáo thuế, chi phí)",
        "examples": [
            ["Quản lý hóa đơn",
             "Nhập hóa đơn thủ công dễ sai số tiền",
             "Thêm kiểm tra/tính tổng tự động khi nhập hóa đơn",
             "Giảm sai sót số liệu", "Cao", "Trung bình"],
            ["Báo cáo thuế",
             "Xuất báo cáo còn thủ công, mất thời gian tổng hợp",
             "Thêm nút xuất báo cáo thuế theo kỳ ra Excel",
             "Tiết kiệm thời gian làm báo cáo", "Trung bình", "Trung bình"],
        ],
    },
    {
        "sheet": "Thu mua",
        "desc": "Thu mua NVL, Mua thiết bị (yêu cầu mua, đánh giá NVL, tiêu chuẩn NVL)",
        "examples": [
            ["Yêu cầu mua hàng",
             "Khó theo dõi yêu cầu đã duyệt/chờ duyệt/đã mua",
             "Thêm trạng thái rõ ràng + lọc theo trạng thái",
             "Theo dõi tiến độ mua hàng dễ hơn", "Cao", "Dễ"],
            ["Đánh giá nguyên vật liệu",
             "So sánh nhà cung cấp còn thủ công",
             "Thêm bảng so sánh điểm đánh giá giữa các NCC",
             "Chọn NCC tốt hơn, minh bạch", "Trung bình", "Trung bình"],
        ],
    },
    {
        "sheet": "Sản xuất",
        "desc": "Quản lý SX, Dữ liệu SX (sản lượng, quy trình), Quản lý kho (nhập/xuất, tồn kho)",
        "examples": [
            ["Quản lý kho - Phiếu nhập/xuất",
             "Nhập liệu phiếu trên tablet ở xưởng còn khó thao tác",
             "Tối ưu form nhập/xuất cho màn hình tablet, nút bấm to hơn",
             "Nhập liệu tại xưởng nhanh, ít sai", "Cao", "Trung bình"],
            ["Dữ liệu sản xuất - Sản lượng",
             "Chưa có cảnh báo khi sản lượng thực tế lệch nhiều so với kế hoạch",
             "Thêm cảnh báo/tô màu khi lệch kế hoạch vượt ngưỡng",
             "Phát hiện bất thường sớm", "Cao", "Trung bình"],
            ["Tồn kho thành phẩm",
             "Khó biết mặt hàng nào sắp hết/tồn lâu",
             "Thêm cảnh báo tồn tối thiểu và hàng tồn lâu ngày",
             "Quản lý tồn kho chủ động", "Trung bình", "Trung bình"],
        ],
    },
    {
        "sheet": "Kỹ thuật",
        "desc": "Đảm bảo & Cải tiến (máy móc, bảo trì, sửa chữa), Phòng phát triển (dự án)",
        "examples": [
            ["Bảo trì máy móc",
             "Dễ quên lịch bảo trì định kỳ",
             "Thêm nhắc nhở lịch bảo trì sắp đến hạn qua thông báo",
             "Bảo trì đúng hạn, giảm hỏng máy", "Cao", "Trung bình"],
            ["Yêu cầu sửa chữa",
             "Khó theo dõi trạng thái xử lý yêu cầu sửa chữa",
             "Thêm dòng thời gian trạng thái cho mỗi yêu cầu",
             "Minh bạch tiến độ sửa chữa", "Trung bình", "Dễ"],
            ["Quản lý dự án",
             "Cập nhật tiến độ dự án còn rời rạc, khó nhìn tổng thể",
             "Thêm biểu đồ tiến độ tổng quan (Gantt) rõ ràng hơn",
             "Nắm tiến độ dự án nhanh", "Trung bình", "Khó"],
        ],
    },
]

N_BLANK_ROWS = 20  # số dòng trống để nhân viên tự điền


def style_department_sheet(ws, dept):
    """Tạo 1 sheet phòng ban với tiêu đề, header, ví dụ mẫu và dòng trống."""
    columns = dept.get("columns", COLUMNS)
    n_cols = len(columns)
    last_col = get_column_letter(n_cols)

    # Đặt độ rộng cột
    for idx, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Hàng 1: tiêu đề phòng ban
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"GÓP Ý CẢI TIẾN ERP  —  {dept['sheet'].upper()}"
    c.font = title_font
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = center
    ws.row_dimensions[1].height = 30

    # Hàng 2: mô tả phạm vi
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = f"Phạm vi: {dept['desc']}   |   Ghi rõ, cụ thể để dễ triển khai. Dòng chữ nghiêng là VÍ DỤ MẪU, có thể xóa."
    c.font = sub_font
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = center
    ws.row_dimensions[2].height = 26

    # Hàng 3: header cột
    header_row = 3
    for idx, (name, _) in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=idx, value=name)
        c.font = head_font
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = center
        c.border = box
    ws.row_dimensions[header_row].height = 34

    # Dữ liệu ví dụ + dòng trống
    examples = dept["examples"]
    start = header_row + 1
    total = len(examples) + N_BLANK_ROWS

    for i in range(total):
        r = start + i
        ws.cell(row=r, column=1, value=i + 1).alignment = center
        is_example = i < len(examples)
        if is_example:
            # cột 2..7 lấy từ ví dụ (Chức năng, Vấn đề, Đề xuất, Lợi ích, Ưu tiên, Độ khó)
            for j, val in enumerate(examples[i], start=2):
                cell = ws.cell(row=r, column=j, value=val)
                cell.font = example_font
                cell.alignment = left if j <= 5 else center
        # Các cột căn giữa: STT + những cột ngắn (ưu tiên, độ khó, ngày...)
        center_cols = {1, 6, 7}
        for ci, (cname, _) in enumerate(columns, start=1):
            if "Ngày" in cname:
                center_cols.add(ci)
        # style toàn bộ ô trong hàng
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = box
            if not is_example:
                cell.font = cell_font
            cell.alignment = center if col in center_cols else left
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=GREY)
        ws.row_dimensions[r].height = 42

    # Data validation dropdown: Mức ưu tiên (cột 6), Độ khó (cột 7)
    end = start + total - 1
    dv_pri = DataValidation(type="list", formula1=PRIORITY, allow_blank=True)
    dv_dif = DataValidation(type="list", formula1=DIFFICULTY, allow_blank=True)
    ws.add_data_validation(dv_pri)
    ws.add_data_validation(dv_dif)
    dv_pri.add(f"F{start}:F{end}")
    dv_dif.add(f"G{start}:G{end}")

    # Freeze header
    ws.freeze_panes = "A4"


def build_guide_sheet(ws):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "HƯỚNG DẪN GÓP Ý CẢI TIẾN HỆ THỐNG ERP — AN BÌNH FOODS"
    c.font = title_font
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = center
    ws.row_dimensions[1].height = 34

    lines = [
        ("", ""),
        ("MỤC ĐÍCH", None),
        ("", "File này để anh/chị ghi lại các LỖI đang gặp và các CẢI TIẾN mong muốn "
             "trên phần mềm ERP. Ý kiến sẽ được tổng hợp và ưu tiên triển khai."),
        ("", ""),
        ("CÁCH ĐIỀN", None),
        ("", "1. Chọn đúng sheet (tab) của phòng ban mình ở phía dưới màn hình."),
        ("", "2. Mỗi dòng là MỘT góp ý. Điền càng cụ thể càng dễ triển khai."),
        ("", "3. Các dòng chữ nghiêng màu xám là VÍ DỤ MẪU — anh/chị có thể xóa hoặc để tham khảo."),
        ("", "4. Cột 'Mức ưu tiên' và 'Độ khó' bấm vào ô để chọn từ danh sách có sẵn."),
        ("", "5. 'Độ khó' nếu không rõ thì chọn 'Chưa rõ' — đội kỹ thuật sẽ đánh giá lại."),
        ("", ""),
        ("Ý NGHĨA CÁC CỘT", None),
        ("", "• Chức năng / màn hình liên quan: tên trang/tính năng đang nói tới (VD: Bảng chấm công, Đơn hàng...)."),
        ("", "• Vấn đề / hạn chế gặp phải: mô tả điều đang gây khó khăn, chậm, hay sai sót."),
        ("", "• Đề xuất cải tiến: anh/chị mong muốn phần mềm làm được gì để tốt hơn."),
        ("", "• Lợi ích mong đợi: cải tiến này giúp ích gì (nhanh hơn, ít lỗi hơn, tiết kiệm thời gian...)."),
        ("", "• Mức ưu tiên: Cao (cần gấp) / Trung bình / Thấp (có cũng tốt)."),
        ("", "• Độ khó: ước tính sơ bộ, có thể để 'Chưa rõ'."),
        ("", "• Người đề xuất, Ngày đề xuất, Ghi chú: thông tin bổ sung."),
        ("", ""),
        ("LƯU Ý", None),
        ("", "Mọi ý kiến đều đáng giá. Không cần lo về kỹ thuật — cứ mô tả điều anh/chị thấy bất tiện, "
             "đội phát triển sẽ tìm cách giải quyết."),
    ]

    r = 2
    for label, text in lines:
        if text is None:  # tiêu đề mục
            ws.merge_cells(f"A{r}:B{r}")
            cell = ws[f"A{r}"]
            cell.value = label
            cell.font = Font(name="Calibri", size=12, bold=True, color=BLUE)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 24
        else:
            ws[f"A{r}"].value = label
            cell = ws[f"B{r}"]
            cell.value = text
            cell.font = cell_font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if text:
                ws.row_dimensions[r].height = 30
        r += 1


def main():
    wb = openpyxl.Workbook()
    guide = wb.active
    guide.title = "Hướng dẫn"
    build_guide_sheet(guide)

    for dept in DEPARTMENTS:
        ws = wb.create_sheet(title=dept["sheet"][:31])
        style_department_sheet(ws, dept)

    wb.save(OUT)
    print("Đã tạo:", OUT)


if __name__ == "__main__":
    main()

